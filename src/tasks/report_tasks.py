import os
from datetime import datetime, timedelta, timezone

from src.celery_app import celery_app
from src.core.database import AsyncSessionLocal
from src.data.repositories.report_repository import ReportRepository
from src.domain.services.batch_service import BatchService
from src.domain.services.product_service import ProductService
from src.domain.services.report_service import ReportService
from src.domain.services.webhook_service import WebhookService
from src.storage.minio_service import minio_service


@celery_app.task(bind=True, max_retries=3)
async def generate_batch_report(
    self, batch_id: int, format: str = "excel", user_email: str | None = None
):
    async with AsyncSessionLocal() as session:
        report_service = ReportService(session)
        batch_service = BatchService(session)
        product_service = ProductService(session)
        webhook_service = WebhookService(session)

        batch = await batch_service.get_by_id(batch_id)
        if batch is None:
            raise ValueError(f"Batch with ID {batch_id} not found")

        report = await report_service.get_last_report(batch_id)
        if report is None:
            raise ValueError(f"Report for batch {batch_id} not found")

        products = await product_service.get_by_batch_id(batch_id)
        file_name = None
        try:
            if format == "excel":
                from openpyxl import Workbook
                from openpyxl.styles import Font
                from openpyxl.utils import get_column_letter

                file_name = f"batch_{batch_id}.xlsx"

                wb = Workbook()
                sheet = wb.active
                sheet.title = "Информация о партии"

                sheet["A1"] = "Номер партии"
                sheet["B1"] = batch.batch_number

                sheet["A2"] = "Дата партии"
                sheet["B2"] = batch.batch_date

                sheet["A3"] = "Статус"
                sheet["B3"] = batch.is_closed

                sheet["A4"] = "Рабочий центр"
                sheet["B4"] = batch.work_center.identifier

                sheet["A5"] = "Смена"
                sheet["B5"] = batch.shift

                sheet["A6"] = "Бригада"
                sheet["B6"] = batch.team

                sheet["A7"] = "Номенклатура"
                sheet["B7"] = batch.nomenclature

                sheet["A8"] = "Начало смены"
                sheet["B8"] = batch.shift_start.replace(tzinfo=None)

                sheet["A9"] = "Окончание смены"
                sheet["B9"] = batch.shift_end.replace(tzinfo=None)

                products_sheet = wb.create_sheet("Продукция")

                products_sheet["A1"] = "ID"
                products_sheet["B1"] = "Уникальный код"
                products_sheet["C1"] = "Аггрегирована"
                products_sheet["D1"] = "Дата аггрегации"

                for i, product in enumerate(products):
                    products_sheet[f"A{i + 2}"] = product.id
                    products_sheet[f"B{i + 2}"] = product.unique_code
                    products_sheet[f"C{i + 2}"] = (
                        "Да" if product.is_aggregated else "Нет"
                    )
                    products_sheet[f"D{i + 2}"] = (
                        product.aggregated_at.replace(tzinfo=None)
                        if product.is_aggregated
                        else "-"
                    )

                stats_sheet = wb.create_sheet("Статистика")
                stats = await batch_service.get_statistics(batch_id)

                stats_sheet["A1"] = "Всего продукции"
                stats_sheet["B1"] = stats["total_products"]

                stats_sheet["A2"] = "Аггрегировано"
                stats_sheet["B2"] = stats["aggregated"]

                stats_sheet["A3"] = "Осталось"
                stats_sheet["B3"] = stats["remaining"]

                stats_sheet["A4"] = "Процент выполнения"
                stats_sheet["B4"] = f"{round(stats['aggregation_rate'] * 100, 2)}%"

                stats_sheet["A5"] = "Средняя скорость"

                end_time = min(datetime.now(timezone.utc), batch.shift_end)

                duration_hours = (end_time - batch.shift_start).total_seconds() / 3600

                average_speed = (
                    round(len(products) / duration_hours, 2)
                    if duration_hours > 0
                    else 0
                )

                stats_sheet["B5"] = f"{average_speed} ед/час"

                for worksheet in wb.worksheets:
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)

                        for cell in column:
                            try:
                                if cell.value is not None:
                                    max_length = max(max_length, len(str(cell.value)))
                            except Exception:
                                pass

                        adjusted_width = max_length + 2
                        worksheet.column_dimensions[
                            column_letter
                        ].width = adjusted_width

                for cell in products_sheet[1]:
                    cell.font = Font(bold=True)

                wb.save(file_name)
            elif format == "pdf":
                from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
                from reportlab.pdfbase import pdfmetrics
                from reportlab.pdfbase.ttfonts import TTFont
                from reportlab.platypus import (
                    PageBreak,
                    Paragraph,
                    SimpleDocTemplate,
                    Spacer,
                )

                file_name = f"batch_{batch.id}.pdf"
                doc = SimpleDocTemplate(file_name)
                pdfmetrics.registerFont(
                    TTFont(
                        "DejaVuSans", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
                    )
                )

                styles = getSampleStyleSheet()

                title_style = ParagraphStyle(
                    "TitleRu", parent=styles["Title"], fontName="DejaVuSans"
                )

                normal_style = ParagraphStyle(
                    "NormalRu", parent=styles["Normal"], fontName="DejaVuSans"
                )

                heading1_style = ParagraphStyle(
                    "HeadingRu", parent=styles["Heading1"], fontName="DejaVuSans"
                )

                heading2_style = ParagraphStyle(
                    "HeadingRu", parent=styles["Heading2"], fontName="DejaVuSans"
                )

                elements = []
                elements.append(Paragraph(f"Партия №{batch.batch_number}", title_style))
                elements.append(Paragraph(f"Дата: {batch.batch_date}", title_style))
                elements.append(
                    Paragraph(
                        f"Рабочий центр: {batch.work_center.identifier}", title_style
                    )
                )
                elements.append(Paragraph(f"Смена: {batch.shift}", title_style))
                elements.append(Paragraph(f"Бригада: {batch.team}", title_style))
                elements.append(Spacer(1, 12))

                stats = await batch_service.get_statistics(batch_id)

                elements.append(Paragraph("Статистика", heading2_style))
                elements.append(
                    Paragraph(
                        f"Всего продукции: {stats['total_products']}", title_style
                    )
                )
                elements.append(
                    Paragraph(f"Агрегировано: {stats['aggregated']}", title_style)
                )
                elements.append(
                    Paragraph(f"Осталось: {stats['remaining']}", title_style)
                )
                elements.append(
                    Paragraph(
                        f"Процент выполнения: "
                        f"{round(stats['aggregation_rate'] * 100, 2)}%",
                        title_style,
                    )
                )

                elements.append(PageBreak())

                elements.append(Paragraph("Продукция", heading1_style))

                for product in products:
                    elements.append(
                        Paragraph(
                            f"{product.unique_code} | "
                            f"{'Да' if product.is_aggregated else 'Нет'}",
                            title_style,
                        )
                    )

                doc.build(elements)

            file_url = minio_service.upload_file(
                bucket="reports", object_name=file_name, file_path=file_name
            )

            expires_at = datetime.now(timezone.utc) + timedelta(days=7)

            report.status = "completed"
            report.file_name = file_name
            report.file_path = file_name
            await session.commit()
            await session.refresh(report)

            await webhook_service.send_event(
                "report_generated",
                {
                    "batch_id": batch.id,
                    "report_type": format,
                    "file_url": file_url,
                    "expires_at": expires_at.isoformat(),
                },
                async_mode=False,
            )
        except Exception:
            report.status = "failed"
            await session.commit()
            await session.refresh(report)
            raise
        finally:
            if file_name and os.path.exists(file_name):
                os.remove(file_name)


@celery_app.task(bind=True, max_retries=3)
async def cleanup_old_files(self):
    try:
        async with AsyncSessionLocal() as session:
            report_repo = ReportRepository(session)

            cutoff_date = datetime.now(timezone.utc) - timedelta(days=30)

            reports = await report_repo.get_older_than(cutoff_date)

            for report in reports:
                try:
                    minio_service.delete_file(
                        bucket="reports", object_name=report.file_name
                    )
                except Exception:
                    pass

                await report_repo.delete(report.id)

            await session.commit()

    except Exception as exc:
        raise self.retry(exc=exc, countdown=60 * (2**self.request.retries))
