import json
import os

from typing import List
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from src.api.v1.schemas.batch import BatchListRequest
from src.domain.services.analytics_service import AnalyticsService
from src.domain.services.product_service import ProductService
from src.domain.services.batch_service import BatchService
from src.core.database import AsyncSessionLocal
from src.core.cache import redis
from src.celery_app import celery_app
from src.storage.minio_service import minio_service

@celery_app.task(bind=True)
async def aggregate_products_batch(self, batch_id: int, codes: List[str]):
    """Массовая агрегация продукции"""

    async with AsyncSessionLocal() as session:
        product_service = ProductService(session)
        batch_service = BatchService(session)

        batch = await batch_service.get_by_id(batch_id)
        if batch is None:
            raise ValueError(f'Batch with ID {batch_id} not found')


        total = len(codes)
        aggregated = 0
        failed = 0
        errors = []

        self.update_state(
            state="PROGRESS",
            meta={
                "current": 0,
                "total": total,
                "percent": 0,
                "aggregated": 0,
                "failed": 0
            }
        )

        for i, code in enumerate(codes):
            try:
                product = await product_service.get_by_unique_code(code, batch_id)
                if product is None:
                    failed += 1
                    errors.append({
                        "code": code,
                        "reason": 'Product not found'
                    })
                    continue
                if product.is_aggregated == True:
                    failed += 1
                    errors.append({
                        "code": code,
                        "reason": 'Product already aggregated'
                    })
                    continue

                await batch_service.aggregate_product(batch_id, code)
                aggregated += 1
            except Exception as e:
                failed += 1
                errors.append({
                    "code": code,
                    "reason": str(e)
                })

            if (i + 1) % 10 == 0 or (i + 1) == total:
                self.update_state(
                    state="PROGRESS",
                    meta={
                        "current": i + 1,
                        "total": total,
                        "percent": round(((i + 1) / total) * 100, 2),
                        "aggregated": aggregated,
                        "failed": failed
                    }
                )
        await session.commit()

        return {
            "success": True,
            "total": total,
            "aggregated": aggregated,
            "failed": failed,
            "errors": errors[:10]
        }


# TODO: import_batches_from_file
@celery_app.task(bind=True, max_retries=1)
def import_batches_from_file(
    self,
    file_url: str,
    user_id: int
):
    """
    Импорт партий из Excel/CSV файла.
    
    Формат файла (Excel):
    | НомерПартии | ДатаПартии | Номенклатура | РабочийЦентр | ... |
    |-------------|------------|--------------|--------------|-----|
    | 22222       | 2024-01-30 | Болт М10     | Цех №1       | ... |
    
    Args:
        file_url: URL файла в MinIO
        user_id: ID пользователя для отправки результата
    
    Returns:
        {
            "success": True,
            "total_rows": 100,
            "created": 95,
            "skipped": 5,
            "errors": [
                {"row": 15, "error": "Duplicate batch number and date"},
                ...
            ]
        }
    """
    pass

# TODO: export_batches_to_file
@celery_app.task
async def export_batches_to_file(
    filters: dict,
    format: str = "excel"
):
    """
    Экспорт списка партий в файл.
    
    Args:
        filters: Фильтры для выборки партий
        format: "excel" или "csv"
    
    Returns:
        {
            "success": True,
            "file_url": "...",
            "total_batches": 150
        }
    """
    print("EXPORT STARTED")
    async with AsyncSessionLocal() as session:
        batch_service = BatchService(session)
        file_name = None
        try:
            filters['offset'] = 0
            filters['limit'] = 100

            batches, total = await batch_service.get_list(BatchListRequest(**filters))
            headers = [
                "ID",
                "Номер партии",
                "Дата партии",
                "Статус",
                "Смена",
                "Бригада",
                "Номенклатура",
                "ЕКН",
            ]

            if format == 'excel':
                file_name = (
                    f"batches_export_"
                    f"{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                    ".xlsx"
                )
                wb = Workbook()
                sheet = wb.active


                
                sheet.append(headers)
                for cell in sheet[1]:
                    cell.font = Font(bold=True)

                for batch in batches:
                    sheet.append([
                        batch.id,
                        batch.batch_number,
                        batch.batch_date,
                        "Закрыта" if batch.is_closed else "Активна",
                        batch.shift,
                        batch.team,
                        batch.nomenclature,
                        batch.ekn_code,
                    ])

                for worksheet in wb.worksheets:
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)

                        for cell in column:
                            try:
                                if cell.value is not None:
                                    max_length = max(
                                        max_length,
                                        len(str(cell.value))
                                    )
                            except Exception:
                                pass

                        adjusted_width = max_length + 2
                        worksheet.column_dimensions[column_letter].width = adjusted_width

                wb.save(file_name)

            elif format == 'csv':
                import csv

                file_name = (
                    f"batches_export_"
                    f"{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                    ".csv"
                )

                with open(file_name, "w", newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)

                    writer.writerow(headers)

                    for batch in batches:
                        writer.writerow([
                            batch.id,
                            batch.batch_number,
                            batch.batch_date,
                            "Закрыта" if batch.is_closed else "Активна",
                            batch.shift,
                            batch.team,
                            batch.nomenclature,
                            batch.ekn_code,
                    ])
            else:
                raise ValueError('Supports only Excel or CSV')
            
            file_url = minio_service.upload_file(
                bucket='exports',
                object_name=file_name,
                file_path=file_name
            )

            return {
                "success": True,
                "file_url": file_url,
                "total_batches": total
            }
        finally:
            if file_name and os.path.exists(file_name):
                os.remove(file_name)


@celery_app.task
async def auto_close_expired_batches():
    async with AsyncSessionLocal() as session:
        batch_service = BatchService(session)
        closed = await batch_service.close_expired_batches()

        return {"closed_batches": closed}


@celery_app.task
async def update_cached_statistics():
    async with AsyncSessionLocal() as session:
        analytic_service = AnalyticsService(session)
        stats = await analytic_service.get_dashboard_statistics()

        stats["cached_at"] = datetime.now(timezone.utc).isoformat()

        await redis.set("dashboard_stats", json.dumps(stats), ex=300)

        return stats

@celery_app.task(bind=True, max_retries=3)
def test_celery_task(self, message: str):
    """
    Тестовая задача для проверки Celery.
    """
    return {"status": "success", "message": message}